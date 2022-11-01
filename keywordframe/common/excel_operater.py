import openpyxl
import os
import time
from keywordframe.config import DirConfig
from openpyxl.styles import Font,PatternFill

class Excel_Operator:
    """
    操作excel文件
    """
    def __init__(self,filename=os.path.join(DirConfig.testdata_dir,"Book1.xlsx")):
        # 获取文件路径
        self.filepath=filename
        # 获取测试用例excel工作簿
        self.wk=openpyxl.load_workbook(filename)

    # 获取工作簿信息
    def get_casesdata(self):

        # h获取excel工作簿所有工作表的数据
        casesname = self.get_casesname()
        self.sheetnames = casesname
        values = []
        # 遍历每个工作表
        for sheetname in self.sheetnames:
            # 获取某个工作表某个区间列的数据
            value = self.get_startcol_endcol_value(sheetname)
            casedata={}
            casedata["casename"]=sheetname
            casedata["stepdata"]=value
            values.append(casedata)
        return values

    # 获取执行用例的工作表名称
    def get_casesname(self,sheetname=DirConfig.casessheetname):
        sheet = self.wk[sheetname]
        casesname =[]
        # 根据是否执行取到用例的名称
        for row in range(2,sheet.max_row+1):
            isexecute = sheet.cell(row,DirConfig.caseIsExcuetecol).value
            if isexecute == 'y':
                casename = sheet.cell(row,DirConfig.caseNamecol).value
                casesname.append(casename)
        return casesname

    def get_startcol_endcol_value(self,sheetname,start=DirConfig.keywordcol,end=DirConfig.actioncol):
        # 获取工作表
        sheet = self.wk[sheetname]
        values = []
        # 遍历对应工作表中的每一行数据
        for row in range(2,sheet.max_row+1):
            stepdata = []
            for col in range(start,end+1):
                value = sheet.cell(row,col).value
                if value != None:
                    stepdata.append(value)
            values.append(stepdata)
        return values

    def write_step_result(self, sheetname,row,col,result):
        """写入测试step的result"""
        casesheet=self.wk[sheetname]
        casesheet.cell(row,col).value = result
        # 颜色填写，green是pass，red是fail
        redfill = PatternFill("solid",fgColor="00FF0000")
        greenfill = PatternFill("solid",fgColor="0000FF00")
        if result == "FALSE":
            # 如果fail当前单元格填red，否则为green
            casesheet.cell(row,col).fill=redfill
        else:
            casesheet.cell(row,col).fill=greenfill
        self.wk.save(self.filepath)

    def write_case_result(self,sheetname=DirConfig.casessheetname,col=DirConfig.caseresultcol):
        sheet = self.wk[sheetname]
        # 颜色填写，green是pass，red是fail
        redfill = PatternFill("solid", fgColor="00FF0000")
        greenfill = PatternFill("solid", fgColor="0000FF00")
        for row in range(2,sheet.max_row+1):
            # 遍历汇总表每一行
            isExecute = sheet.cell(row,DirConfig.caseIsExcuetecol).value
            if isExecute == "y":
                # 获取用例名称
                casename = sheet.cell(row,DirConfig.casestepnamecol).value
                # 获取该用例所有测试step的结果
                stepresults = self.get_sheet_col_values(casename)
                if "FALSE" in stepresults:
                    # 写入汇总结果:FALSE
                    sheet.cell(row.col).fill=redfill
                    sheet.cell(row,col).value = "FALSE"
                else:
                    # 写入汇总结果：PASS
                    sheet.cell(row,col).fill=greenfill
                    sheet.cell(row,col).value = "PASS"


    def get_sheet_col_values(self,sheetname):
        values=[]
        sheet = self.wk[sheetname]
        for row in range(2,sheet.max_row+1):
            stepresult = sheet.cell(row,DirConfig.casestepresult).value
            values.append(stepresult)
        return values

    # 写入测试step执行的时间
    def write_step_time(self,sheetname,row,col=DirConfig.actionTimecol):
        sheet = self.wk[sheetname]
        current_time = time.strftime("%Y-%m-%d %H:%M:%S",time.localtime())
        sheet.cell(row,col).value=current_time
        self.wk.save(self.filepath)

if __name__ == '__main__':
    print(Excel_Operator().get_startcol_endcol_value("普通用户登陆"))
